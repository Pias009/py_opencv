import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


def _autosize(ws, min_width=10, max_width=50):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def generate_report_xlsx(entry, output_path):
    """Build an Excel (.xlsx) summary of a single vehicle-count run.

    entry: same dict saved to data/results/<id>.json.
    """
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    when = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(entry.get("started_at") or time.time()))
    duration = entry.get("duration_sec")
    duration_str = f"{duration:.1f} s" if duration is not None else "-"

    ws["A1"] = "Vehicle Count Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generated: {when}"
    ws["A2"].font = Font(italic=True, color="666666")

    rows = [
        ("Video file", entry.get("video", "-")),
        ("Status", entry.get("status", "-")),
        ("Total frames", entry.get("total_frames", "-")),
        ("Processing time", duration_str),
        ("Total vehicles counted", entry.get("count", 0)),
        ("Model used", entry.get("model_used", "-")),
    ]
    r = 4
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=2, value=value)
        r += 1
    _autosize(ws)

    # --- Per-line / direction sheet ---
    lines = entry.get("lines") or {}
    if lines:
        ws2 = wb.create_sheet("By Road-Direction")
        headers = ["Line / Road", "In", "Out", "Total"]
        for c, h in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        r = 2
        for name, v in lines.items():
            in_c, out_c = v.get("in", 0), v.get("out", 0)
            ws2.cell(row=r, column=1, value=name)
            ws2.cell(row=r, column=2, value=in_c)
            ws2.cell(row=r, column=3, value=out_c)
            ws2.cell(row=r, column=4, value=in_c + out_c)
            r += 1
        _autosize(ws2)
    else:
        incoming, outgoing = entry.get("incoming", 0), entry.get("outgoing", 0)
        if incoming or outgoing:
            ws2 = wb.create_sheet("By Direction")
            ws2.cell(row=1, column=1, value="Direction").fill = HEADER_FILL
            ws2.cell(row=1, column=1).font = HEADER_FONT
            ws2.cell(row=1, column=2, value="Count").fill = HEADER_FILL
            ws2.cell(row=1, column=2).font = HEADER_FONT
            ws2.cell(row=2, column=1, value="Incoming")
            ws2.cell(row=2, column=2, value=incoming)
            ws2.cell(row=3, column=1, value="Outgoing")
            ws2.cell(row=3, column=2, value=outgoing)
            _autosize(ws2)

    # --- Per-category sheet ---
    categories = entry.get("categories") or {}
    if categories:
        ws3 = wb.create_sheet("By Vehicle Type")
        headers = ["Category", "Count", "Share (%)"]
        for c, h in enumerate(headers, start=1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        total = sum(categories.values()) or 1
        r = 2
        for name, n in sorted(categories.items(), key=lambda kv: -kv[1]):
            ws3.cell(row=r, column=1, value=name)
            ws3.cell(row=r, column=2, value=n)
            ws3.cell(row=r, column=3, value=round(100 * n / total, 1))
            r += 1
        _autosize(ws3)
        note_row = r + 1
        ws3.cell(row=note_row, column=1,
                 value=("Note: vehicle types are classified using an automated detection model. "
                        "Types not covered by the model are reported as 'Other' or their closest match."))
        ws3.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="808080")

    wb.save(output_path)
    return output_path
